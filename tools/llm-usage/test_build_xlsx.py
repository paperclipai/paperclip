#!/usr/bin/env python3
"""Tests der Excel-Mappe. Aufruf: python3 -m pytest test_build_xlsx.py -q

`query` wird ersetzt, damit die Mappe ohne laufende Postgres-Instanz baubar
ist — geprueft wird das Format, nicht die Abfrage.
"""
from datetime import date

from openpyxl import Workbook

import build_xlsx
import query

# `tag` als `date`, so wie psycopg2 die `::date`-Spalte liefert.
TAG = date(2026, 8, 22)
PER_LLM_PER_DAY = [
    # (model, tag, calls, in_tok, cached_tok, out_tok, tokens, dur_sec)
    ("gemma4-31b-it", TAG, 300, 900_000, 0, 100_000, 1_000_000, 3600),
    ("claude-sonnet-4-6", TAG, 50, 400_000, 0, 100_000, 500_000, 1800),
]
AGENT_HOUR = [
    # (agent, tag, stunde, model, calls, tokens)
    ("CTO", TAG, "09:00", "gemma4-31b-it", 12, 40_000),
    ("CEO", TAG, "09:00", "claude-sonnet-4-6", 3, 20_000),
]


def _stub(monkeypatch):
    monkeypatch.setattr(query, "per_llm_per_day", lambda days=7: PER_LLM_PER_DAY)
    monkeypatch.setattr(query, "agent_hour", lambda days=7: AGENT_HOUR)
    monkeypatch.setattr(
        query, "matrix_day_by_model",
        lambda days=7: (["2026-08-22"], ["gemma4-31b-it"],
                        {("2026-08-22", "gemma4-31b-it"): 300}),
    )


def test_alter_tag_bekommt_den_damaligen_ort(monkeypatch):
    """Vor dem 06.07.2026 gab es nur den Mac Studio — die Mappe blickt sieben
    Tage zurueck und darf einen alten Tag nicht mit heutigen Geraeten fuellen."""
    alt = [("gemma-4-31b-it-mlx", date(2026, 5, 20), 10, 100, 0, 10, 110, 60)]
    monkeypatch.setattr(query, "per_llm_per_day", lambda days=7: alt)
    monkeypatch.setattr(
        query, "matrix_day_by_model",
        lambda days=7: (["2026-05-20"], ["gemma-4-31b-it-mlx"], {}),
    )
    wb = Workbook()
    build_xlsx.sheet_llm_per_day(wb, 7)
    zeilen = _zeilen(wb.active)
    zeile = next(z for z in zeilen if z and z[0] == "gemma-4-31b-it-mlx")
    assert zeile[1] == "Mac Studio"


def _zeilen(ws):
    return [[c.value for c in row] for row in ws.iter_rows()]


SB = {
    "katalog": {"gemma4-31b-it": {"quant": "Q8_0", "ctx": 98304}},
    "denken": {"gemma4-31b-it": [1000, 12]},
}


def test_sheet_llm_pro_tag_traegt_den_steckbrief(monkeypatch):
    _stub(monkeypatch)
    wb = Workbook()
    build_xlsx.sheet_llm_per_day(wb, 7, sb=SB)
    zeilen = _zeilen(wb.active)
    kopf = next(z for z in zeilen if z and z[0] == "LLM")
    for spalte in ("Quant", "CTX", "Thinking", "Denkquote"):
        assert spalte in kopf, spalte
    zeile = next(z for z in zeilen if z and z[0] == "gemma4-31b-it")
    werte = dict(zip(kopf, zeile))
    assert werte["Quant"] == "Q8_0"
    assert werte["CTX"] == "96K"
    assert werte["Thinking"] == "off (1 %)"
    # Die Quote als Zahl, nicht als Text — sonst kann Excel nicht damit rechnen.
    assert werte["Denkquote"] == 0.012


def test_denkquote_bleibt_leer_wenn_nicht_messbar(monkeypatch):
    _stub(monkeypatch)
    wb = Workbook()
    build_xlsx.sheet_llm_per_day(wb, 7, sb=SB)
    zeilen = _zeilen(wb.active)
    kopf = next(z for z in zeilen if z and z[0] == "LLM")
    zeile = next(z for z in zeilen if z and z[0] == "claude-sonnet-4-6")
    assert dict(zip(kopf, zeile))["Denkquote"] is None


def test_sheet_agent_je_stunde_traegt_den_steckbrief(monkeypatch):
    _stub(monkeypatch)
    wb = Workbook()
    build_xlsx.sheet_agent_hour(wb, 7, sb=SB)
    zeilen = _zeilen(wb["Agent je Stunde"])
    kopf = next(z for z in zeilen if z and z[0] == "Agent")
    zeile = next(z for z in zeilen if z and z[0] == "CTO")
    werte = dict(zip(kopf, zeile))
    assert (werte["Quant"], werte["CTX"], werte["Thinking"]) == \
        ("Q8_0", "96K", "off (1 %)")


def test_sheet_llm_pro_tag_hat_eine_wo_spalte(monkeypatch):
    _stub(monkeypatch)
    wb = Workbook()
    build_xlsx.sheet_llm_per_day(wb, 7)
    zeilen = _zeilen(wb.active)
    kopf = next(z for z in zeilen if z and z[0] == "LLM")
    assert "Wo" in kopf
    spalte = kopf.index("Wo")
    daten = {z[0]: z[spalte] for z in zeilen if z and z[0] in
             ("gemma4-31b-it", "claude-sonnet-4-6")}
    assert daten == {"gemma4-31b-it": "RTX", "claude-sonnet-4-6": "Cloud"}


def test_sheet_agent_je_stunde_hat_eine_wo_spalte(monkeypatch):
    _stub(monkeypatch)
    wb = Workbook()
    build_xlsx.sheet_agent_hour(wb, 7)
    zeilen = _zeilen(wb["Agent je Stunde"])
    kopf = next(z for z in zeilen if z and z[0] == "Agent")
    assert "Wo" in kopf
    spalte = kopf.index("Wo")
    orte = [z[spalte] for z in zeilen if z and z[0] in ("CTO", "CEO")]
    assert orte == ["RTX", "Cloud"]
