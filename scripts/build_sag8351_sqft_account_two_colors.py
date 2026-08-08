#!/usr/bin/env python3
"""Build the approved SAG-8742 Quartz correction from a read-only SQLite source."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sqlite3
import zipfile
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
DB = ROOT / "data/2026-08-05/sag-8483/sage_extract.db"
START, END = "2023-08-04", "2026-08-04"
PACKAGE_FILES = (
    "Sage_Quartz_Corrected_Parent_Color_Lane.xlsx", "annual_parent_year_color_lane.csv",
    "parent_company_lineage.csv", "retail_catalog_color_crossover.csv",
    "house_po_sku_manufacturer.csv", "coverage.csv", "query_receipt.sql",
    "validation.json", "README.md",
)
ANNUAL_FIELDS = ("year", "lane", "parent_company_id", "parent_company_name", "material_product_id", "material_color", "source_line_count", "calculated_measure")
LINEAGE_FIELDS = ("lane", "source_company_id", "source_company_name", "parent_company_id", "parent_company_name", "parent_is_source_company", "source_line_count")
CROSSOVER_FIELDS = ("material_product_id", "material_color", "retail_product_id", "retail_sku", "retail_catalog_id", "retail_catalog_code", "retail_catalog_name", "exact_consumer_description", "normalized_comparison_key", "candidate_count", "mapping_status")
MANUFACTURER_FIELDS = ("source_product_id", "source_sku", "mfr_id", "manufacturer_company_id", "manufacturer_company_name", "affected_po_line_count", "calculated_po_measure", "mapping_status")
COVERAGE_FIELDS = ("lane", "coverage_scope", "exclusion_reason", "source_line_count", "source_native_quantity", "calculated_measure")
DETAIL_FIELDS = ("year", "lane", "parent_company_id", "parent_company_name", "source_company_id", "source_company_name", "material_product_id", "material_color", "calculated_measure")


def decimal(value: object) -> Decimal:
    return Decimal(str(value if value not in (None, "") else 0))


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def color(row: sqlite3.Row) -> str:
    return clean(row["ColorDescription"]) or clean(row["ConsumerDescription"]) or clean(row["Description"]) or "(blank material color)"


def read_only(db_path: Path) -> sqlite3.Connection:
    connection = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    connection.row_factory = sqlite3.Row
    return connection


LANE_SQL = {
    "RO": """
      SELECT 'RO' AS lane, r.ROProductID AS source_line_id, ro.RODate AS source_date,
             r.DealerQty AS native_quantity,
             COALESCE(rp.SquareFootage, h.SquareFootage) AS reportability_square_footage,
             CASE WHEN rp.MaterialID = 2510 THEN rp.ProductID ELSE h.ProductID END AS material_product_id,
             CASE WHEN rp.MaterialID = 2510 THEN rp.ColorDescription ELSE h.ColorDescription END AS ColorDescription,
             CASE WHEN rp.MaterialID = 2510 THEN rp.ConsumerDescription ELSE h.ConsumerDescription END AS ConsumerDescription,
             CASE WHEN rp.MaterialID = 2510 THEN rp.Description ELSE h.Description END AS Description,
             CASE WHEN rp.MaterialID = 2510 THEN rp.MFRID ELSE h.MFRID END AS MFRID,
             ro.DealerCompanyID AS source_company_id, sc.CompanyName AS source_company_name,
             COALESCE(sc.ParentCompanyID, sc.CompanyID) AS parent_company_id,
             COALESCE(pc.CompanyName, sc.CompanyName, '[UNASSIGNED COMPANY]') AS parent_company_name
      FROM ORD_ROProduct AS r
      JOIN ORD_RetailOrder AS ro ON ro.RetailOrderID = r.RetailOrderID
      LEFT JOIN CAT_Product AS rp ON rp.ProductID = r.ProductID
      LEFT JOIN CAT_Product AS h ON h.ProductID = rp.XRefProductID
        AND h.ProductCatalogID = 1 AND h.MaterialID = 2510
      LEFT JOIN COM_Company AS sc ON sc.CompanyID = ro.DealerCompanyID
      LEFT JOIN COM_Company AS pc ON pc.CompanyID = COALESCE(sc.ParentCompanyID, sc.CompanyID)
      WHERE date(ro.RODate) >= ? AND date(ro.RODate) < ? AND ro.OrderStatusID <> 40
        AND (rp.MaterialID = 2510 OR h.ProductID IS NOT NULL)
    """,
    "MO": """
      SELECT 'MO' AS lane, m.MOProductID AS source_line_id, mo.MODate AS source_date,
             m.Quantity AS native_quantity, mp.SquareFootage AS reportability_square_footage,
             mp.ProductID AS material_product_id, mp.ColorDescription, mp.ConsumerDescription, mp.Description, mp.MFRID,
             mo.CompanyID AS source_company_id, sc.CompanyName AS source_company_name,
             COALESCE(sc.ParentCompanyID, sc.CompanyID) AS parent_company_id,
             COALESCE(pc.CompanyName, sc.CompanyName, '[UNASSIGNED COMPANY]') AS parent_company_name
      FROM ORD_MOProduct AS m
      JOIN ORD_MaterialOrder AS mo ON mo.MaterialOrderID = m.MaterialOrderID
      JOIN CAT_Product AS mp ON mp.ProductID = m.ProductID AND mp.MaterialID = 2510
      LEFT JOIN COM_Company AS sc ON sc.CompanyID = mo.CompanyID
      LEFT JOIN COM_Company AS pc ON pc.CompanyID = COALESCE(sc.ParentCompanyID, sc.CompanyID)
      WHERE date(mo.MODate) >= ? AND date(mo.MODate) < ? AND mo.OrderStatusID <> 40
    """,
    "PO": """
      SELECT 'PO' AS lane, p.POProductID AS source_line_id, po.CreatedOn AS source_date,
             p.QtyOrdered AS native_quantity, mp.SquareFootage AS reportability_square_footage,
             mp.ProductID AS material_product_id, mp.ColorDescription, mp.ConsumerDescription, mp.Description, mp.MFRID, mp.SKU,
             mo.CompanyID AS source_company_id, sc.CompanyName AS source_company_name,
             COALESCE(sc.ParentCompanyID, sc.CompanyID) AS parent_company_id,
             COALESCE(pc.CompanyName, sc.CompanyName, '[UNASSIGNED COMPANY]') AS parent_company_name
      FROM ORD_POProduct AS p
      JOIN ORD_PurchaseOrder AS po ON po.PurchaseOrderID = p.PurchaseOrderID
      LEFT JOIN ORD_MaterialOrder AS mo ON mo.MaterialOrderID = po.MaterialOrderID
      JOIN CAT_Product AS mp ON mp.ProductID = p.ProductID AND mp.MaterialID = 2510
      LEFT JOIN COM_Company AS sc ON sc.CompanyID = mo.CompanyID
      LEFT JOIN COM_Company AS pc ON pc.CompanyID = COALESCE(sc.ParentCompanyID, sc.CompanyID)
      WHERE date(po.CreatedOn) >= ? AND date(po.CreatedOn) < ? AND po.OrderStatusID <> 40
    """,
}


def classify(lane: str, row: sqlite3.Row) -> tuple[bool, Decimal, str | None]:
    quantity, square_footage = decimal(row["native_quantity"]), decimal(row["reportability_square_footage"])
    if lane == "RO":
        return quantity > 0 and square_footage > 0, quantity, None if quantity > 0 and square_footage > 0 else "Non-positive DealerQty or reportability SquareFootage"
    measure = quantity * square_footage
    if quantity <= 0:
        return False, measure, "Non-positive Quantity"
    if square_footage <= 0:
        return False, measure, "Non-positive SquareFootage"
    return True, measure, None


def collect(db_path: Path = DB) -> tuple[list[dict], list[dict], list[dict], list[dict], list[dict], list[dict]]:
    """Read source rows once for the generator; all access is SQLite ``mode=ro``."""
    connection = read_only(db_path)
    try:
        raw_by_lane = {lane: list(connection.execute(sql, (START, END))) for lane, sql in LANE_SQL.items()}
        detail: list[dict] = []
        coverage: list[dict] = []
        lineage_counts: dict[tuple, int] = defaultdict(int)
        annual_totals: dict[tuple, dict] = defaultdict(lambda: {"source_line_count": 0, "calculated_measure": Decimal(0)})
        for lane, rows in raw_by_lane.items():
            cohort_native = sum((decimal(row["native_quantity"]) for row in rows), Decimal(0))
            cohort_measure = sum((classify(lane, row)[1] for row in rows), Decimal(0))
            reportable_native = reportable_measure = Decimal(0)
            reportable_count = 0
            exclusions: dict[str, dict] = defaultdict(lambda: {"source_line_count": 0, "source_native_quantity": Decimal(0), "calculated_measure": Decimal(0)})
            missing_lineage = 0
            for row in rows:
                valid, measure, reason = classify(lane, row)
                quantity = decimal(row["native_quantity"])
                if not row["source_company_id"]:
                    missing_lineage += 1
                if not valid:
                    excluded = exclusions[reason or "Excluded"]
                    excluded["source_line_count"] += 1
                    excluded["source_native_quantity"] += quantity
                    excluded["calculated_measure"] += measure
                    continue
                reportable_count += 1
                reportable_native += quantity
                reportable_measure += measure
                stage = {
                    "year": clean(row["source_date"])[:4], "lane": lane,
                    "parent_company_id": row["parent_company_id"], "parent_company_name": clean(row["parent_company_name"]),
                    "source_company_id": row["source_company_id"], "source_company_name": clean(row["source_company_name"]),
                    "material_product_id": row["material_product_id"], "material_color": color(row), "calculated_measure": measure,
                }
                detail.append(stage)
                key = tuple(stage[field] for field in ANNUAL_FIELDS[:6])
                annual_totals[key]["source_line_count"] += 1
                annual_totals[key]["calculated_measure"] += measure
                lineage_key = (lane, stage["source_company_id"], stage["source_company_name"], stage["parent_company_id"], stage["parent_company_name"], stage["source_company_id"] == stage["parent_company_id"])
                lineage_counts[lineage_key] += 1
            coverage.extend([
                {"lane": lane, "coverage_scope": "Cohort", "exclusion_reason": "All cohort", "source_line_count": len(rows), "source_native_quantity": cohort_native, "calculated_measure": cohort_measure},
                {"lane": lane, "coverage_scope": "Reportable", "exclusion_reason": "All reportable", "source_line_count": reportable_count, "source_native_quantity": reportable_native, "calculated_measure": reportable_measure},
            ])
            coverage.extend({"lane": lane, "coverage_scope": "Excluded", "exclusion_reason": reason, **values} for reason, values in sorted(exclusions.items()))
            if missing_lineage:
                coverage.append({"lane": lane, "coverage_scope": "Informational", "exclusion_reason": "Missing company lineage", "source_line_count": missing_lineage, "source_native_quantity": Decimal(0), "calculated_measure": Decimal(0)})

        annual = [{**dict(zip(ANNUAL_FIELDS[:6], key)), **value} for key, value in sorted(annual_totals.items(), key=lambda item: tuple("" if value is None else str(value) for value in item[0]))]
        lineage = [{**dict(zip(LINEAGE_FIELDS[:6], key)), "source_line_count": count} for key, count in sorted(lineage_counts.items(), key=str)]
        materials = list(connection.execute("SELECT ProductID, ColorDescription, ConsumerDescription, Description FROM CAT_Product WHERE MaterialID = 2510 ORDER BY ProductID"))
        candidate_rows = list(connection.execute("""
          SELECT r.XRefProductID AS material_product_id, m.ColorDescription, m.ConsumerDescription AS material_consumer_description, m.Description AS material_description,
                 r.ProductID AS retail_product_id, r.SKU AS retail_sku, r.ProductCatalogID AS retail_catalog_id,
                 c.ProductCatalogCode AS retail_catalog_code, c.Description AS retail_catalog_name,
                 r.ConsumerDescription AS exact_consumer_description
          FROM CAT_Product AS r
          JOIN CAT_Product AS m ON m.ProductID = r.XRefProductID AND m.MaterialID = 2510
          JOIN CAT_ProductCatalog AS c ON c.ProductCatalogID = r.ProductCatalogID
          WHERE r.XRefProductID IS NOT NULL AND COALESCE(r.Inactive, 0) = 0 AND COALESCE(c.Inactive, 0) = 0
            AND TRIM(COALESCE(r.ConsumerDescription, '')) <> ''
          GROUP BY r.XRefProductID, r.ProductID, r.ProductCatalogID, r.ConsumerDescription
          ORDER BY r.XRefProductID, r.ProductID, r.ProductCatalogID, r.ConsumerDescription
        """))
        candidates: dict[int, list[sqlite3.Row]] = defaultdict(list)
        for row in candidate_rows:
            candidates[row["material_product_id"]].append(row)
        crossover: list[dict] = []
        for material in materials:
            rows = candidates[material["ProductID"]]
            count = len(rows)
            status = "Unmatched" if count == 0 else "Matched" if count == 1 else "Ambiguous"
            if not rows:
                crossover.append({"material_product_id": material["ProductID"], "material_color": clean(material["ColorDescription"]) or clean(material["ConsumerDescription"]) or clean(material["Description"]) or "(blank material color)", "retail_product_id": None, "retail_sku": None, "retail_catalog_id": None, "retail_catalog_code": None, "retail_catalog_name": None, "exact_consumer_description": None, "normalized_comparison_key": None, "candidate_count": 0, "mapping_status": status})
            for row in rows:
                raw = row["exact_consumer_description"]
                crossover.append({"material_product_id": row["material_product_id"], "material_color": clean(row["ColorDescription"]) or clean(row["material_consumer_description"]) or clean(row["material_description"]) or "(blank material color)", "retail_product_id": row["retail_product_id"], "retail_sku": row["retail_sku"], "retail_catalog_id": row["retail_catalog_id"], "retail_catalog_code": row["retail_catalog_code"], "retail_catalog_name": row["retail_catalog_name"], "exact_consumer_description": raw, "normalized_comparison_key": clean(raw).casefold(), "candidate_count": count, "mapping_status": status})
        company_rows: dict[object, list[sqlite3.Row]] = defaultdict(list)
        for company in connection.execute("SELECT CompanyID, CompanyName FROM COM_Company ORDER BY CompanyID, CompanyName"):
            company_rows[company["CompanyID"]].append(company)
        mfr_totals: dict[tuple, dict] = defaultdict(lambda: {"affected_po_line_count": 0, "calculated_po_measure": Decimal(0)})
        for row in raw_by_lane["PO"]:
            valid, measure, _ = classify("PO", row)
            source_product_id, sku, mfr_id = row["material_product_id"], clean(row["SKU"]), row["MFRID"]
            matches = [] if mfr_id is None else company_rows[mfr_id]
            if mfr_id is None:
                relations = [(None, None, "Null MFRID")]
            elif not matches:
                relations = [(None, None, "Orphan MFRID")]
            elif len(matches) == 1:
                relations = [(matches[0]["CompanyID"], clean(matches[0]["CompanyName"]), "Matched")]
            else:
                relations = [(company["CompanyID"], clean(company["CompanyName"]), "Multiple COM_Company matches") for company in matches]
            for company_id, company_name, status in relations:
                item = mfr_totals[(source_product_id, sku, mfr_id, company_id, company_name, status)]
                item["affected_po_line_count"] += 1
                item["calculated_po_measure"] += measure if valid else Decimal(0)
        manufacturer = [{"source_product_id": key[0], "source_sku": key[1], "mfr_id": key[2], "manufacturer_company_id": key[3], "manufacturer_company_name": key[4], **value, "mapping_status": key[5]} for key, value in sorted(mfr_totals.items(), key=str)]
        return detail, annual, lineage, crossover, manufacturer, coverage
    finally:
        connection.close()


def write_csv(path: Path, fields: tuple[str, ...], rows: list[dict]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: str(row.get(field, "")) if isinstance(row.get(field), Decimal) else row.get(field, "") for field in fields})


def title(field: str) -> str:
    return field.replace("_", " ").title().replace("Id", "ID").replace("Sku", "SKU")


def write_table(sheet, fields: tuple[str, ...], rows: list[dict]) -> None:
    sheet.append([title(field) for field in fields])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17365D")
    for row in rows:
        sheet.append([str(row[field]) if isinstance(row.get(field), Decimal) else row.get(field) for field in fields])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def write_workbook(path: Path, detail: list[dict], annual: list[dict], lineage: list[dict], crossover: list[dict], manufacturer: list[dict], coverage: list[dict]) -> None:
    workbook = Workbook()
    detail_sheet = workbook.active
    detail_sheet.title = "Parent Color Year Detail"
    write_table(detail_sheet, DETAIL_FIELDS, detail)
    rollup_sheet = workbook.create_sheet("Annual Parent Rollup")
    write_table(rollup_sheet, ANNUAL_FIELDS[:-1], annual)
    rollup_sheet.cell(1, len(ANNUAL_FIELDS)).value = title(ANNUAL_FIELDS[-1])
    rollup_sheet.cell(1, len(ANNUAL_FIELDS)).font = Font(bold=True, color="FFFFFF")
    rollup_sheet.cell(1, len(ANNUAL_FIELDS)).fill = PatternFill("solid", fgColor="17365D")
    for number, row in enumerate(annual, start=2):
        for column, field in enumerate(ANNUAL_FIELDS[:-1], start=1):
            rollup_sheet.cell(number, column).value = row[field]
        rollup_sheet.cell(number, len(ANNUAL_FIELDS)).value = f'=SUMIFS(\'Parent Color Year Detail\'!$I:$I,\'Parent Color Year Detail\'!$A:$A,A{number},\'Parent Color Year Detail\'!$B:$B,B{number},\'Parent Color Year Detail\'!$C:$C,C{number},\'Parent Color Year Detail\'!$G:$G,E{number},\'Parent Color Year Detail\'!$H:$H,F{number})'
    write_table(workbook.create_sheet("Parent Company Lineage"), LINEAGE_FIELDS, lineage)
    write_table(workbook.create_sheet("Retail Catalog Crossover"), CROSSOVER_FIELDS, crossover)
    write_table(workbook.create_sheet("House PO SKU Manufacturer"), MANUFACTURER_FIELDS, manufacturer)
    write_table(workbook.create_sheet("Coverage"), COVERAGE_FIELDS, coverage)
    methodology = workbook.create_sheet("Methodology")
    methodology.append(["Methodology"])
    methodology.append(["RO is Decimal(ORD_ROProduct.DealerQty), reportable only when DealerQty and resolved source-or-house SquareFootage are positive."])
    methodology.append(["MO is Decimal(Quantity) * Decimal(CAT_Product.SquareFootage); PO is Decimal(QtyOrdered) * Decimal(CAT_Product.SquareFootage). Lanes are separate and must not be added together."])
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = min(48, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
    workbook.save(path)


def receipt_sql() -> str:
    return f"""-- SAG-8742 corrected Quartz query receipt; source opened by the scripts with SQLite mode=ro.
-- Source: data/2026-08-05/sag-8483/sage_extract.db
-- Window: [{START}, {END}); cancellation predicate: OrderStatusID <> 40.
-- RO cohort and classification (DealerQty remains the measure; SquareFootage is only reportability):
WITH ro_cohort AS (
 SELECT r.ROProductID source_line_id, r.DealerQty dealer_qty, COALESCE(rp.SquareFootage,h.SquareFootage) reportability_square_footage
 FROM ORD_ROProduct r JOIN ORD_RetailOrder ro ON ro.RetailOrderID=r.RetailOrderID
 LEFT JOIN CAT_Product rp ON rp.ProductID=r.ProductID
 LEFT JOIN CAT_Product h ON h.ProductID=rp.XRefProductID AND h.ProductCatalogID=1 AND h.MaterialID=2510
 WHERE date(ro.RODate)>='{START}' AND date(ro.RODate)<'{END}' AND ro.OrderStatusID<>40
   AND (rp.MaterialID=2510 OR h.ProductID IS NOT NULL)
) SELECT COUNT(*), SUM(CAST(dealer_qty AS NUMERIC)), SUM(CASE WHEN CAST(dealer_qty AS REAL)>0 AND CAST(reportability_square_footage AS REAL)>0 THEN 1 ELSE 0 END) FROM ro_cohort;
-- MO: ORD_MOProduct.Quantity * CAT_Product.SquareFootage. PO: ORD_POProduct.QtyOrdered * CAT_Product.SquareFootage.
-- Parent key: COALESCE(source_company.ParentCompanyID, source_company.CompanyID).
-- Crossover grain: (material_product_id, retail_product_id, retail_catalog_id, exact_consumer_description).
-- Manufacturer mapping: CAT_Product.MFRID LEFT JOIN COM_Company.
"""


def package(output_dir: Path, package_path: Path) -> None:
    package_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(package_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name in sorted(PACKAGE_FILES):
            archive.write(output_dir / name, name)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, default=ROOT / "artifacts/2026-08-07/sag-8742")
    parser.add_argument("--package-path", type=Path, default=ROOT / "artifacts/SAG-8742_quartz_analysis_corrected_deliverable.zip")
    parser.add_argument("--db-path", type=Path, default=DB)
    args = parser.parse_args()
    if not args.db_path.is_file():
        raise SystemExit(f"Missing source SQLite extract: {args.db_path}")
    args.output_dir.mkdir(parents=True, exist_ok=True)
    detail, annual, lineage, crossover, manufacturer, coverage = collect(args.db_path)
    write_workbook(args.output_dir / PACKAGE_FILES[0], detail, annual, lineage, crossover, manufacturer, coverage)
    write_csv(args.output_dir / "annual_parent_year_color_lane.csv", ANNUAL_FIELDS, annual)
    write_csv(args.output_dir / "parent_company_lineage.csv", LINEAGE_FIELDS, lineage)
    write_csv(args.output_dir / "retail_catalog_color_crossover.csv", CROSSOVER_FIELDS, crossover)
    write_csv(args.output_dir / "house_po_sku_manufacturer.csv", MANUFACTURER_FIELDS, manufacturer)
    write_csv(args.output_dir / "coverage.csv", COVERAGE_FIELDS, coverage)
    (args.output_dir / "query_receipt.sql").write_text(receipt_sql(), encoding="utf-8")
    (args.output_dir / "README.md").write_text("# Sage Quartz corrected parent/color/lane analysis\n\nRO is `Decimal(ORD_ROProduct.DealerQty)` only. MO is `Decimal(Quantity) * Decimal(CAT_Product.SquareFootage)` and PO is `Decimal(QtyOrdered) * Decimal(CAT_Product.SquareFootage)`. These lanes are separate business measures and must not be added together. The RO coverage sheet retains both cohort and excluded controls; reportability requires positive DealerQty and positive resolved source-or-house SquareFootage.\n", encoding="utf-8")
    producer_validation = {
        "status": "generator checksum manifest awaiting independent verifier",
        "source_path": str(args.db_path.relative_to(ROOT)),
        "window": f"[{START}, {END})",
        "lane_totals": {lane: str(sum((row["calculated_measure"] for row in annual if row["lane"] == lane), Decimal(0))) for lane in ("RO", "MO", "PO")},
        "package_member_names": sorted(PACKAGE_FILES),
        "checksum_scope": "all package members except validation.json",
        "artifact_sha256": {name: hashlib.sha256((args.output_dir / name).read_bytes()).hexdigest() for name in PACKAGE_FILES if name != "validation.json"},
    }
    (args.output_dir / "validation.json").write_text(json.dumps(producer_validation, indent=2) + "\n", encoding="utf-8")
    package(args.output_dir, args.package_path)
    print(json.dumps({"output_dir": str(args.output_dir), "package": str(args.package_path), "reportable_lane_totals": producer_validation["lane_totals"], "ro_coverage": [row for row in coverage if row["lane"] == "RO"]}, default=str, indent=2))


if __name__ == "__main__":
    main()
