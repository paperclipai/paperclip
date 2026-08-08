#!/usr/bin/env python3
"""Independent read-only source replay and package verifier for SAG-8742."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sqlite3
import zipfile
from collections import Counter, defaultdict
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DB = ROOT / "data/2026-08-05/sag-8483/sage_extract.db"
START, END = "2023-08-04", "2026-08-04"
EXPECTED_CONTROLS = {
    "RO": {"cohort_count": 58056, "cohort_total": Decimal("2557013.560"), "reportable_count": 48098, "reportable_total": Decimal("2061862.510"), "excluded_count": 9958, "excluded_total": Decimal("495151.050")},
    "MO": {"cohort_count": 41942, "cohort_total": Decimal("1974481.800"), "reportable_count": 26577, "reportable_total": Decimal("1974481.800"), "excluded_count": 15365, "excluded_total": Decimal("0")},
    "PO": {"cohort_count": 40298, "cohort_total": Decimal("1893200.390"), "reportable_count": 25767, "reportable_total": Decimal("1893200.390"), "excluded_count": 14531, "excluded_total": Decimal("0")},
}
PACKAGE_FILES = {"Sage_Quartz_Corrected_Parent_Color_Lane.xlsx", "annual_parent_year_color_lane.csv", "parent_company_lineage.csv", "retail_catalog_color_crossover.csv", "house_po_sku_manufacturer.csv", "coverage.csv", "query_receipt.sql", "validation.json", "README.md"}
SHEETS = ["Parent Color Year Detail", "Annual Parent Rollup", "Parent Company Lineage", "Retail Catalog Crossover", "House PO SKU Manufacturer", "Coverage", "Methodology"]


def decimal(value: object) -> Decimal:
    return Decimal(str(value if value not in (None, "") else 0))


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def color(row: sqlite3.Row) -> str:
    return clean(row["ColorDescription"]) or clean(row["ConsumerDescription"]) or clean(row["Description"]) or "(blank material color)"


def open_source() -> sqlite3.Connection:
    connection = sqlite3.connect(f"file:{DB}?mode=ro", uri=True)
    connection.row_factory = sqlite3.Row
    return connection


# Deliberately separate from the generator: this SQL is the verifier's own replay contract.
SQL = {
    "RO": """
      SELECT r.ROProductID source_line_id, ro.RODate source_date, r.DealerQty native_quantity,
             COALESCE(rp.SquareFootage,h.SquareFootage) reportability_square_footage,
             CASE WHEN rp.MaterialID=2510 THEN rp.ProductID ELSE h.ProductID END material_product_id,
             CASE WHEN rp.MaterialID=2510 THEN rp.ColorDescription ELSE h.ColorDescription END ColorDescription,
             CASE WHEN rp.MaterialID=2510 THEN rp.ConsumerDescription ELSE h.ConsumerDescription END ConsumerDescription,
             CASE WHEN rp.MaterialID=2510 THEN rp.Description ELSE h.Description END Description,
             ro.DealerCompanyID source_company_id, sc.CompanyName source_company_name,
             COALESCE(sc.ParentCompanyID,sc.CompanyID) parent_company_id,
             COALESCE(pc.CompanyName,sc.CompanyName,'[UNASSIGNED COMPANY]') parent_company_name
      FROM ORD_ROProduct r JOIN ORD_RetailOrder ro ON ro.RetailOrderID=r.RetailOrderID
      LEFT JOIN CAT_Product rp ON rp.ProductID=r.ProductID
      LEFT JOIN CAT_Product h ON h.ProductID=rp.XRefProductID AND h.ProductCatalogID=1 AND h.MaterialID=2510
      LEFT JOIN COM_Company sc ON sc.CompanyID=ro.DealerCompanyID
      LEFT JOIN COM_Company pc ON pc.CompanyID=COALESCE(sc.ParentCompanyID,sc.CompanyID)
      WHERE date(ro.RODate)>=? AND date(ro.RODate)<? AND ro.OrderStatusID<>40
        AND (rp.MaterialID=2510 OR h.ProductID IS NOT NULL)
    """,
    "MO": """
      SELECT m.MOProductID source_line_id, mo.MODate source_date, m.Quantity native_quantity, mp.SquareFootage reportability_square_footage,
             mp.ProductID material_product_id, mp.ColorDescription, mp.ConsumerDescription, mp.Description,
             mo.CompanyID source_company_id, sc.CompanyName source_company_name, COALESCE(sc.ParentCompanyID,sc.CompanyID) parent_company_id,
             COALESCE(pc.CompanyName,sc.CompanyName,'[UNASSIGNED COMPANY]') parent_company_name
      FROM ORD_MOProduct m JOIN ORD_MaterialOrder mo ON mo.MaterialOrderID=m.MaterialOrderID
      JOIN CAT_Product mp ON mp.ProductID=m.ProductID AND mp.MaterialID=2510
      LEFT JOIN COM_Company sc ON sc.CompanyID=mo.CompanyID LEFT JOIN COM_Company pc ON pc.CompanyID=COALESCE(sc.ParentCompanyID,sc.CompanyID)
      WHERE date(mo.MODate)>=? AND date(mo.MODate)<? AND mo.OrderStatusID<>40
    """,
    "PO": """
      SELECT p.POProductID source_line_id, po.CreatedOn source_date, p.QtyOrdered native_quantity, mp.SquareFootage reportability_square_footage,
             mp.ProductID material_product_id, mp.ColorDescription, mp.ConsumerDescription, mp.Description, mp.MFRID, mp.SKU,
             mp.SKU source_sku, mp.MFRID mfr_id,
             mo.CompanyID source_company_id, sc.CompanyName source_company_name, COALESCE(sc.ParentCompanyID,sc.CompanyID) parent_company_id,
             COALESCE(pc.CompanyName,sc.CompanyName,'[UNASSIGNED COMPANY]') parent_company_name
      FROM ORD_POProduct p JOIN ORD_PurchaseOrder po ON po.PurchaseOrderID=p.PurchaseOrderID
      LEFT JOIN ORD_MaterialOrder mo ON mo.MaterialOrderID=po.MaterialOrderID
      JOIN CAT_Product mp ON mp.ProductID=p.ProductID AND mp.MaterialID=2510
      LEFT JOIN COM_Company sc ON sc.CompanyID=mo.CompanyID LEFT JOIN COM_Company pc ON pc.CompanyID=COALESCE(sc.ParentCompanyID,sc.CompanyID)
      WHERE date(po.CreatedOn)>=? AND date(po.CreatedOn)<? AND po.OrderStatusID<>40
    """,
}


def classified(lane: str, row: sqlite3.Row) -> tuple[bool, Decimal]:
    quantity, square_footage = decimal(row["native_quantity"]), decimal(row["reportability_square_footage"])
    if lane == "RO":
        return quantity > 0 and square_footage > 0, quantity
    return quantity > 0 and square_footage > 0, quantity * square_footage


def exclusion_reason(lane: str, row: sqlite3.Row) -> str | None:
    quantity, square_footage = decimal(row["native_quantity"]), decimal(row["reportability_square_footage"])
    if lane == "RO":
        return None if quantity > 0 and square_footage > 0 else "Non-positive DealerQty or reportability SquareFootage"
    if quantity <= 0:
        return "Non-positive Quantity"
    if square_footage <= 0:
        return "Non-positive SquareFootage"
    return None


def replay(connection: sqlite3.Connection) -> dict[str, dict]:
    results: dict[str, dict] = {}
    for lane, query in SQL.items():
        rows = list(connection.execute(query, (START, END)))
        reportable = []
        aggregates: dict[tuple, dict] = defaultdict(lambda: {"source_line_count": 0, "calculated_measure": Decimal(0)})
        lineage: dict[tuple, int] = defaultdict(int)
        coverage: dict[tuple, dict] = defaultdict(lambda: {"source_line_count": 0, "source_native_quantity": Decimal(0), "calculated_measure": Decimal(0)})
        reportable_native_quantity = Decimal(0)
        missing_lineage = 0
        for row in rows:
            valid, measure = classified(lane, row)
            quantity = decimal(row["native_quantity"])
            if not row["source_company_id"]:
                missing_lineage += 1
            if valid:
                reportable.append((row, measure))
                key = (clean(row["source_date"])[:4], lane, row["parent_company_id"], clean(row["parent_company_name"]), row["material_product_id"], color(row))
                aggregates[key]["source_line_count"] += 1
                aggregates[key]["calculated_measure"] += measure
                lineage[(lane, row["source_company_id"], clean(row["source_company_name"]), row["parent_company_id"], clean(row["parent_company_name"]), row["source_company_id"] == row["parent_company_id"])] += 1
                reportable_native_quantity += quantity
            else:
                excluded = coverage[(lane, "Excluded", exclusion_reason(lane, row) or "Excluded")]
                excluded["source_line_count"] += 1
                excluded["source_native_quantity"] += quantity
                excluded["calculated_measure"] += measure
        cohort_total = sum((classified(lane, row)[1] for row in rows), Decimal(0))
        reportable_total = sum((measure for _, measure in reportable), Decimal(0))
        coverage[(lane, "Cohort", "All cohort")] = {"source_line_count": len(rows), "source_native_quantity": sum((decimal(row["native_quantity"]) for row in rows), Decimal(0)), "calculated_measure": cohort_total}
        coverage[(lane, "Reportable", "All reportable")] = {"source_line_count": len(reportable), "source_native_quantity": reportable_native_quantity, "calculated_measure": reportable_total}
        if missing_lineage:
            coverage[(lane, "Informational", "Missing company lineage")] = {"source_line_count": missing_lineage, "source_native_quantity": Decimal(0), "calculated_measure": Decimal(0)}
        results[lane] = {"rows": rows, "reportable": reportable, "aggregates": aggregates, "lineage": lineage, "coverage": coverage, "cohort_count": len(rows), "cohort_total": cohort_total, "reportable_count": len(reportable), "reportable_total": reportable_total}
        results[lane]["excluded_count"] = results[lane]["cohort_count"] - results[lane]["reportable_count"]
        results[lane]["excluded_total"] = results[lane]["cohort_total"] - results[lane]["reportable_total"]
    return results


def expected_lineage(source: dict[str, dict]) -> dict[tuple, int]:
    expected: dict[tuple, int] = defaultdict(int)
    for lane, data in source.items():
        for row, _ in data["reportable"]:
            expected[(lane, nullable_int(row["source_company_id"]), clean(row["source_company_name"]),
                      nullable_int(row["parent_company_id"]), clean(row["parent_company_name"]),
                      nullable_int(row["source_company_id"]) == nullable_int(row["parent_company_id"]))] += 1
    return expected


def expected_manufacturer(connection: sqlite3.Connection, source: dict[str, dict]) -> dict[tuple, tuple[int, Decimal]]:
    company_rows: dict[object, list[tuple[object, str]]] = defaultdict(list)
    for company in connection.execute("SELECT CompanyID, CompanyName FROM COM_Company ORDER BY CompanyID, CompanyName"):
        company_rows[company[0]].append((company[0], clean(company[1])))
    expected: dict[tuple, tuple[int, Decimal]] = {}
    for row in source["PO"]["rows"]:
        mfr_id = row["mfr_id"]
        matches = [] if mfr_id is None else company_rows[mfr_id]
        if mfr_id is None:
            relations = [(None, None, "Null MFRID")]
        elif not matches:
            relations = [(None, None, "Orphan MFRID")]
        elif len(matches) == 1:
            relations = [(matches[0][0], matches[0][1], "Matched")]
        else:
            relations = [(company_id, name, "Multiple COM_Company matches") for company_id, name in matches]
        valid, measure = classified("PO", row)
        for company_id, company_name, status in relations:
            key = (nullable_int(row["material_product_id"]), clean(row["source_sku"]), nullable_int(mfr_id), company_id, company_name, status)
            count, total = expected.get(key, (0, Decimal(0)))
            expected[key] = (count + 1, total + (measure if valid else Decimal(0)))
    return expected


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def nullable_int(value: str) -> int | None:
    return int(value) if value not in (None, "") else None


def independent_candidates(connection: sqlite3.Connection) -> tuple[list[tuple], set[int]]:
    materials = {row[0] for row in connection.execute("SELECT ProductID FROM CAT_Product WHERE MaterialID=2510")}
    candidates = [tuple(row) for row in connection.execute("""
      SELECT r.XRefProductID, r.ProductID, r.ProductCatalogID, r.ConsumerDescription
      FROM CAT_Product r JOIN CAT_ProductCatalog c ON c.ProductCatalogID=r.ProductCatalogID
      JOIN CAT_Product m ON m.ProductID=r.XRefProductID AND m.MaterialID=2510
      WHERE r.XRefProductID IS NOT NULL AND COALESCE(r.Inactive,0)=0 AND COALESCE(c.Inactive,0)=0
        AND TRIM(COALESCE(r.ConsumerDescription,''))<>''
      GROUP BY r.XRefProductID,r.ProductID,r.ProductCatalogID,r.ConsumerDescription
      ORDER BY r.XRefProductID,r.ProductID,r.ProductCatalogID,r.ConsumerDescription
    """)]
    return candidates, materials


def validate_workbook(path: Path, annual: list[dict[str, str]]) -> int:
    workbook = load_workbook(path, read_only=False, data_only=False)
    if workbook.sheetnames != SHEETS:
        raise SystemExit(f"workbook sheets differ: {workbook.sheetnames}")
    detail_headers = [cell.value for cell in workbook["Parent Color Year Detail"][1]]
    for required in ("Lane", "Parent Company ID", "Parent Company Name", "Source Company ID", "Source Company Name", "Calculated Measure"):
        if required not in detail_headers:
            raise SystemExit(f"workbook detail header missing {required}")
    summary = workbook["Annual Parent Rollup"]
    if summary.max_row != len(annual) + 1:
        raise SystemExit("workbook annual summary row count differs from annual CSV")
    if any(not isinstance(summary.cell(row, 8).value, str) or not summary.cell(row, 8).value.startswith("=SUMIFS(") for row in range(2, summary.max_row + 1)):
        raise SystemExit("workbook annual summary lacks required detail-derived SUMIFS formulas")
    error_tokens = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
    errors = sum(1 for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row if isinstance(cell.value, str) and any(token in cell.value for token in error_tokens))
    if errors:
        raise SystemExit(f"workbook contains {errors} formula error token(s)")
    return errors


def verify_package(output_dir: Path, package_path: Path, validation: dict) -> None:
    if not package_path.is_file():
        raise SystemExit(f"package path does not exist: {package_path}")
    persisted_hashes = validation.get("artifact_sha256")
    expected_names = sorted(PACKAGE_FILES)
    if (not isinstance(persisted_hashes, dict)
            or validation.get("package_member_names") != expected_names
            or validation.get("checksum_scope") != "all package members except validation.json"
            or set(persisted_hashes) != PACKAGE_FILES - {"validation.json"}):
        raise SystemExit("validation.json lacks canonical member/checksum evidence")
    for name in PACKAGE_FILES - {"validation.json"}:
        actual = hashlib.sha256((output_dir / name).read_bytes()).hexdigest()
        if persisted_hashes.get(name) != actual:
            raise SystemExit(f"persisted checksum mismatch for {name}")
    with zipfile.ZipFile(package_path) as archive:
        names = archive.namelist()
        if len(names) != len(set(names)) or names != expected_names or archive.testzip() is not None:
            raise SystemExit("corrected ZIP membership or CRC check failed")
        for name in expected_names:
            if archive.read(name) != (output_dir / name).read_bytes():
                raise SystemExit(f"ZIP member bytes differ from output file: {name}")


def verify(args: argparse.Namespace) -> dict:
    output_dir, package_path = args.output_dir, args.package_path
    if not DB.is_file():
        raise SystemExit(f"approved source SQLite extract is missing: {DB}")
    missing = [name for name in PACKAGE_FILES if not (output_dir / name).is_file()]
    if missing:
        raise SystemExit(f"output directory lacks required package members: {sorted(missing)}")
    receipt = (output_dir / "query_receipt.sql").read_text(encoding="utf-8")
    if "COALESCE(rp.SquareFootage,h.SquareFootage)" not in receipt or "DealerQty dealer_qty" not in receipt or "DealerQty *" in receipt:
        raise SystemExit("receipt fails the approved RO native-measure/reportability contract")
    connection = open_source()
    try:
        source = replay(connection)
        source_candidates, materials = independent_candidates(connection)
    finally:
        connection.close()
    controls = {lane: {key: source[lane][key] for key in EXPECTED_CONTROLS[lane]} for lane in EXPECTED_CONTROLS}
    if controls != EXPECTED_CONTROLS:
        raise SystemExit(f"lane benchmark mismatch: {controls} != {EXPECTED_CONTROLS}")
    annual = read_csv(output_dir / "annual_parent_year_color_lane.csv")
    expected_annual = {key: total for lane in source.values() for key, total in lane["aggregates"].items()}
    actual_annual: dict[tuple, dict] = {}
    for row in annual:
        key = (row["year"], row["lane"], nullable_int(row["parent_company_id"]), row["parent_company_name"], int(row["material_product_id"]), row["material_color"])
        if key in actual_annual:
            raise SystemExit(f"duplicate annual parent key: {key}")
        actual_annual[key] = {"source_line_count": int(row["source_line_count"]), "calculated_measure": decimal(row["calculated_measure"])}
    if actual_annual != expected_annual:
        raise SystemExit(f"annual source replay mismatch (missing={len(set(expected_annual)-set(actual_annual))}, extra={len(set(actual_annual)-set(expected_annual))})")
    coverage = read_csv(output_dir / "coverage.csv")
    expected_coverage = {key: total for lane in source.values() for key, total in lane["coverage"].items()}
    actual_coverage: dict[tuple, dict] = {}
    for row in coverage:
        key = (row["lane"], row["coverage_scope"], row["exclusion_reason"])
        if key in actual_coverage:
            raise SystemExit(f"duplicate coverage control: {key}")
        actual_coverage[key] = {"source_line_count": int(row["source_line_count"]), "source_native_quantity": decimal(row["source_native_quantity"]), "calculated_measure": decimal(row["calculated_measure"])}
    if actual_coverage != expected_coverage:
        raise SystemExit(f"coverage source replay mismatch (missing={len(set(expected_coverage)-set(actual_coverage))}, extra={len(set(actual_coverage)-set(expected_coverage))})")
    lineage = read_csv(output_dir / "parent_company_lineage.csv")
    actual_lineage: dict[tuple, int] = {}
    for row in lineage:
        key = (row["lane"], nullable_int(row["source_company_id"]), row["source_company_name"], nullable_int(row["parent_company_id"]), row["parent_company_name"], row["parent_is_source_company"] == "True")
        if row["parent_is_source_company"] not in ("True", "False") or key in actual_lineage:
            raise SystemExit("parent lineage rows lack a unique fallback/identity key")
        actual_lineage[key] = int(row["source_line_count"])
    if actual_lineage != expected_lineage(source):
        raise SystemExit("parent lineage does not match independent COALESCE(source_company.ParentCompanyID, source_company.CompanyID) replay")
    crossover = read_csv(output_dir / "retail_catalog_color_crossover.csv")
    emitted_candidates = sorted((int(row["material_product_id"]), int(row["retail_product_id"]), int(row["retail_catalog_id"]), row["exact_consumer_description"]) for row in crossover if int(row["candidate_count"]) > 0)
    if emitted_candidates != source_candidates:
        raise SystemExit("lossless retail candidate relation differs from independent source replay")
    counts = Counter(material for material, *_ in source_candidates)
    for row in crossover:
        count = int(row["candidate_count"])
        expected_status = "Unmatched" if count == 0 else "Matched" if count == 1 else "Ambiguous"
        if row["mapping_status"] != expected_status or (count == 0 and (row["retail_product_id"] or row["exact_consumer_description"])):
            raise SystemExit("crossover cardinality/status contract failed")
    unmatched_materials = {int(row["material_product_id"]) for row in crossover if int(row["candidate_count"]) == 0}
    if unmatched_materials != materials - set(counts):
        raise SystemExit("crossover unmatched material sentinel relation differs from source")
    manufacturer = read_csv(output_dir / "house_po_sku_manufacturer.csv")
    actual_manufacturer = {(nullable_int(row["source_product_id"]), row["source_sku"], nullable_int(row["mfr_id"]), nullable_int(row["manufacturer_company_id"]), row["manufacturer_company_name"] or None, row["mapping_status"]): (int(row["affected_po_line_count"]), decimal(row["calculated_po_measure"])) for row in manufacturer}
    connection = open_source()
    try:
        expected_mfr = expected_manufacturer(connection, source)
    finally:
        connection.close()
    if actual_manufacturer != expected_mfr:
        raise SystemExit("manufacturer mapping does not match independent CAT_Product.MFRID -> COM_Company replay")
    errors = validate_workbook(output_dir / "Sage_Quartz_Corrected_Parent_Color_Lane.xlsx", annual)
    validation = json.loads((output_dir / "validation.json").read_text(encoding="utf-8"))
    verify_package(output_dir, package_path, validation)
    validation = {
        "status": "independent verifier passed", "source_path": str(DB.relative_to(ROOT)), "source_sha256": hashlib.sha256(DB.read_bytes()).hexdigest(), "window": f"[{START}, {END})",
        "ro_controls": {key: str(value) for key, value in controls["RO"].items()},
        "lane_replay": {lane: {"source_cohort_line_count": data["cohort_count"], "source_reportable_line_count": data["reportable_count"], "source_reportable_total": str(data["reportable_total"]), "generated_annual_total": str(sum((total["calculated_measure"] for total in data["aggregates"].values()), Decimal(0))), "difference": str(data["reportable_total"] - sum((total["calculated_measure"] for total in data["aggregates"].values()), Decimal(0)))} for lane, data in source.items()},
        "parent_rollup_rows": len(annual), "parent_lineage_rows": len(lineage), "crossover_source_candidates": len(source_candidates), "crossover_emitted_candidates": len(emitted_candidates), "crossover_status_counts": dict(Counter(row["mapping_status"] for row in crossover)), "manufacturer_status_counts": dict(Counter(row["mapping_status"] for row in manufacturer)), "workbook_formula_error_tokens": errors,
        "artifact_sha256": {name: hashlib.sha256((output_dir / name).read_bytes()).hexdigest() for name in PACKAGE_FILES if name != "validation.json"},
        "package_member_names": sorted(PACKAGE_FILES), "verification_command": "python3 scripts/verify_sag8351_sqft_account_two_colors.py --output-dir artifacts/2026-08-07/sag-8742 --package-path artifacts/SAG-8742_quartz_analysis_corrected_deliverable.zip",
    }
    validation["package_sha256"] = hashlib.sha256(package_path.read_bytes()).hexdigest()
    print(json.dumps(validation, sort_keys=True))
    return validation


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, default=ROOT / "artifacts/2026-08-07/sag-8742")
    parser.add_argument("--package-path", type=Path, default=ROOT / "artifacts/SAG-8742_quartz_analysis_corrected_deliverable.zip")
    verify(parser.parse_args())


if __name__ == "__main__":
    main()
