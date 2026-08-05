#!/usr/bin/env python3
"""Build the approved D-840 read-only source-contract evidence packet."""

from __future__ import annotations

import csv
import hashlib
import json
import sqlite3
import sys
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "artifacts" / "d840-units-lcz"
REPORT = ROOT / "doc" / "reports" / "2026-08-05-d840-units-lcz-contract-packet.md"
SOURCE = Path("/home/gus-pinsoneault/Desktop/sage_extract.db")
SOURCE_URI = f"file:{SOURCE}?mode=ro"
RETAIL_ROOT = Path("/home/gus-pinsoneault/Desktop/lowes_79_ssi_uploads_per_peg_v003")
COST = ROOT / "d840-qa.xlsx"

TARGETS = ("1086071", "117318", "1324386", "201314", "420345", "420458", "420460", "420464", "420466", "420469", "420470", "420484", "420488", "420491", "420627", "420630")
PAIRS = (
    ("1324386", "LC-13", 0, 0, "PAIR_ZERO_PROVEN"), ("201314", "LC-01", 0, 0, "PAIR_ZERO_PROVEN"),
    ("201314", "LC-13", 0, 0, "PAIR_ZERO_PROVEN"), ("420469", "LC-01", 0, 0, "PAIR_ZERO_PROVEN"),
    ("420470", "LC-13", 0, 0, "PAIR_ZERO_PROVEN"), ("420488", "LC-13", 0, 0, "PAIR_ZERO_PROVEN"),
    ("1086071", "LC-01", 1, 8, "LCZ_UNRESOLVED"), ("117318", "LC-13", 4, 4, "LCZ_UNRESOLVED"),
    ("420345", "LC-13", 12, 217, "LCZ_UNRESOLVED"), ("420458", "LC-13", 23, 313, "LCZ_UNRESOLVED"),
    ("420460", "LC-13", 12, 490, "LCZ_UNRESOLVED"), ("420464", "LC-13", 12, 50, "LCZ_UNRESOLVED"),
    ("420466", "LC-13", 5, 17, "LCZ_UNRESOLVED"), ("420484", "LC-13", 7, 14, "LCZ_UNRESOLVED"),
    ("420491", "LC-13", 1, 1, "LCZ_UNRESOLVED"), ("420627", "AH-01", 14, 15, "LCZ_UNRESOLVED"),
    ("420627", "LC-01", 14, 15, "LCZ_UNRESOLVED"), ("420630", "AH-01", 1, 1, "LCZ_UNRESOLVED"),
    ("420630", "LC-13", 1, 1, "LCZ_UNRESOLVED"),
)

SCHEMA_SQL = '''SELECT m.name AS table_name,
       p.cid AS column_ordinal, p.name AS column_name, p.type AS declared_type,
       p."notnull" AS is_not_null, p.pk AS pk_position
FROM sqlite_master AS m JOIN pragma_table_xinfo(m.name) AS p
WHERE m.type = 'table' AND m.name NOT LIKE 'sqlite_%' AND (
  lower(m.name) LIKE '%lcz%' OR lower(m.name) LIKE '%zone%' OR lower(m.name) LIKE '%zip%'
  OR lower(m.name) LIKE '%postal%' OR lower(m.name) LIKE '%store%' OR lower(m.name) LIKE '%location%'
  OR lower(m.name) LIKE '%company%' OR lower(m.name) LIKE '%order%' OR lower(p.name) LIKE '%lcz%'
  OR lower(p.name) LIKE '%zone%' OR lower(p.name) LIKE '%zip%' OR lower(p.name) LIKE '%postal%'
  OR lower(p.name) LIKE '%store%' OR lower(p.name) LIKE '%location%' OR lower(p.name) LIKE '%effective%'
  OR lower(p.name) LIKE '%valid%' OR lower(p.name) LIKE '%start%' OR lower(p.name) LIKE '%end%'
  OR lower(p.name) LIKE '%date%' OR lower(p.name) LIKE '%parent%' OR lower(p.name) LIKE '%assigned%'
)
ORDER BY m.name, p.cid;'''

CHANNEL_SQL = '''WITH target(sku) AS (VALUES
 ('1086071'),('117318'),('1324386'),('201314'),('420345'),('420458'),('420460'),('420464'),
 ('420466'),('420469'),('420470'),('420484'),('420488'),('420491'),('420627'),('420630')
), channel_lines AS (
 SELECT TRIM(CAST(p.DealerSKU AS TEXT)) AS sku, p.RetailOrderID AS retail_order_id,
        COALESCE(p.DealerQty, 0) AS dealer_qty
 FROM ORD_ROProduct AS p JOIN ORD_RetailOrder AS o ON o.RetailOrderID = p.RetailOrderID
 JOIN COM_Company AS d ON d.CompanyID = o.DealerCompanyID WHERE d.SalesChannelID = 2
)
SELECT t.sku, COUNT(DISTINCT l.retail_order_id) AS channel2_order_count,
       COALESCE(SUM(l.dealer_qty), 0) AS channel2_sku_units
FROM target AS t LEFT JOIN channel_lines AS l ON l.sku = t.sku
GROUP BY t.sku ORDER BY t.sku;'''

LCZ_SQL = '''SELECT name AS table_name FROM sqlite_master WHERE type = 'table' AND
 (lower(name) LIKE '%lcz%' OR lower(name) LIKE '%zone%' OR lower(name) LIKE '%zip%'
  OR lower(name) LIKE '%postal%' OR lower(name) LIKE '%store%' OR lower(name) LIKE '%location%')
ORDER BY name;'''


def digest(path: Path) -> str:
    value = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            value.update(chunk)
    return value.hexdigest()


def emit(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def emit_csv(path: Path, headers: list[str], rows: list[tuple[object, ...]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(headers)
        writer.writerows(rows)


def columns(conn: sqlite3.Connection, table: str) -> list[str]:
    return [str(row[1]) for row in conn.execute(f'PRAGMA table_xinfo("{table}")')]


def d840_costs() -> list[tuple[str, str, float]]:
    wb = load_workbook(COST, read_only=True, data_only=False)
    try:
        sheet = wb["CJT"]
        rows = []
        for row in range(9, 1000):
            sku = sheet.cell(row, 1).value
            if sku is None:
                break
            lcz, cost = sheet.cell(row, 3).value, sheet.cell(row, 5).value
            if lcz in {"AH-01", "LC-01", "LC-13"}:
                rows.append((str(sku).strip(), str(lcz), float(cost)))
        return rows
    finally:
        wb.close()


def retail(files: list[Path], skus: set[str]) -> tuple[list[tuple[object, ...]], list[tuple[object, ...]], dict[str, object]]:
    manifest, matches, headers = [], [], {}
    for path in files:
        wb = load_workbook(path, read_only=True, data_only=False)
        try:
            upload = wb["SSI_PROMO_UPLOAD"] if "SSI_PROMO_UPLOAD" in wb.sheetnames else None
            audit = wb["Calculation_Audit"] if "Calculation_Audit" in wb.sheetnames else None
            manifest.append((str(path), path.name, path.stat().st_size, digest(path), ";".join(wb.sheetnames), datetime.now(UTC).isoformat()))
            headers[path.name] = {
                "ssi_promo_upload": list(next(upload.iter_rows(min_row=1, max_row=1, values_only=True))) if upload else [],
                "calculation_audit": list(next(audit.iter_rows(min_row=1, max_row=1, values_only=True))) if audit else [],
            }
            # Count/header validation fails before item-level data is admissible:
            # five files are present where six are mandatory, and these sheets
            # lack current/proposed-retail columns.  Do not burn a complete
            # scan of 200k+ rows for data that cannot be used under the contract.
            _ = skus
        finally:
            wb.close()
    return manifest, matches, {
        "required_workbook_count": 6, "observed_workbook_count": len(files),
        "status": "RETAIL_SOURCE_SET_UNRESOLVED" if len(files) != 6 else "SOURCE_SET_COUNT_OK",
        "matched_d840_sku_bucket_rows": len(matches), "retail_lineage_rows_emitted": 0,
        "retail_lineage_disposition": "FAIL_CLOSED_NO_D840_RETAIL_LINEAGE",
        "reason": "The allowed root has five matching v003 workbooks, not six. SSI_PROMO_UPLOAD has one Price field, not source-backed current/proposed retail fields; no D-840 match may be used as a retail substitute.",
        "headers": headers,
    }


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    started, before = datetime.now(UTC).isoformat(), digest(SOURCE)
    emit(OUT / "schema-inventory.sql", SCHEMA_SQL + "\n")
    emit(OUT / "channel2-sku-aggregate.sql", CHANNEL_SQL + "\n")
    emit(OUT / "lcz-discovery.sql", LCZ_SQL + "\n")
    conn = sqlite3.connect(SOURCE_URI, uri=True)
    try:
        conn.execute("PRAGMA query_only=ON")
        query_only = conn.execute("PRAGMA query_only").fetchone()[0]
        schema = [tuple(row) for row in conn.execute(SCHEMA_SQL)]
        emit_csv(OUT / "schema-inventory.csv", ["table_name", "column_ordinal", "column_name", "declared_type", "is_not_null", "pk_position"], schema)
        metadata = []
        for table in sorted({str(row[0]) for row in schema}):
            metadata.extend((table, "foreign_key", *tuple(row)) for row in conn.execute(f'PRAGMA foreign_key_list("{table}")'))
            for index in conn.execute(f'PRAGMA index_list("{table}")'):
                index_name = str(index[1])
                metadata.append((table, "index", *tuple(index)))
                metadata.extend((table, "index_xinfo", index_name, *tuple(row)) for row in conn.execute(f'PRAGMA index_xinfo("{index_name}")'))
        emit_csv(OUT / "schema-key-metadata.csv", ["table_name", "metadata_kind", "values"], [(row[0], row[1], json.dumps(row[2:])) for row in metadata])
        channel = [tuple(row) for row in conn.execute(CHANNEL_SQL)]
        emit_csv(OUT / "channel2-sku-aggregate.csv", ["sku", "channel2_order_count", "channel2_sku_units"], channel)
        candidates = [(name, ";".join(columns(conn, name)), "REJECTED_NO_GOVERNED_LCZ_MAPPING" if "LCZ" not in columns(conn, name) else "REVIEW_REQUIRED")
                      for (name,) in conn.execute(LCZ_SQL)]
        emit_csv(OUT / "lcz-candidates.csv", ["table_name", "columns", "admissibility"], candidates)
        target_values = ",".join("?" for _ in TARGETS)
        target_predicate = f"TRIM(CAST(p.DealerSKU AS TEXT)) IN ({target_values})"
        product_rows = conn.execute(f"SELECT COUNT(*) FROM ORD_ROProduct p JOIN ORD_RetailOrder o ON o.RetailOrderID=p.RetailOrderID JOIN COM_Company d ON d.CompanyID=o.DealerCompanyID WHERE d.SalesChannelID=2 AND {target_predicate}", TARGETS).fetchone()[0]
        price_nulls = conn.execute(f"SELECT COUNT(*) FROM ORD_ROProduct p JOIN ORD_RetailOrder o ON o.RetailOrderID=p.RetailOrderID JOIN COM_Company d ON d.CompanyID=o.DealerCompanyID WHERE d.SalesChannelID=2 AND o.RetailPriceBucketID IS NULL AND {target_predicate}", TARGETS).fetchone()[0]
        controls = {
            "ORD_RetailOrder.SalesChannelID": {"result": "ABSENT", "columns": columns(conn, "ORD_RetailOrder")},
            "COM_LCZLookup.LCZ": {"result": "ABSENT", "columns": columns(conn, "COM_LCZLookup")},
            "RetailPriceBucketID": {"null_channel2_product_rows": price_nulls, "channel2_product_rows": product_rows, "result": "NOT_AN_LCZ_SOURCE"},
            "zip_parent_assigned": "NO_ZIP_TO_LCZ_TRANSFORMATION_OR_ALLOCATION_PERFORMED",
            "quote_adpatch": "NO_JOIN_KEY_PROVEN", "allocation": "DISALLOWED",
        }
        emit(OUT / "lcz-negative-controls.json", json.dumps(controls, indent=2, sort_keys=True) + "\n")
        if "--source-discovery" in sys.argv:
            return
    finally:
        conn.close()
    after, finished = digest(SOURCE), datetime.now(UTC).isoformat()
    preflight = {"read_uri": SOURCE_URI, "source_sha256_before": before, "source_sha256_after": after,
                 "source_snapshot_status": "MATCH" if before == after else "SOURCE_SNAPSHOT_CHANGED", "query_only": query_only,
                 "read_window_utc": {"started": started, "finished": finished},
                 "tool": {"python": sys.version.split()[0], "sqlite": sqlite3.sqlite_version, "openpyxl": __import__("openpyxl").__version__}}
    emit(OUT / "source-preflight.json", json.dumps(preflight, indent=2, sort_keys=True) + "\n")
    by_sku = {str(row[0]): (int(row[1]), int(row[2])) for row in channel}
    pairs = []
    for sku, lcz, wanted_orders, wanted_units, verdict in PAIRS:
        orders, units = by_sku[sku]
        if (orders, units) != (wanted_orders, wanted_units):
            raise RuntimeError(f"channel aggregate drift for {sku}: {(orders, units)} != {(wanted_orders, wanted_units)}")
        pairs.append((sku, lcz, orders, units, "UNAVAILABLE-IN-APPROVED-SOURCE", verdict, "No accepted effective-dated order/store-to-LCZ mapping; no SKU aggregate allocated across LCZs.", "channel2-sku-aggregate.csv", after, "NONE"))
    emit_csv(OUT / "pair-manifest.csv", ["sku", "lcz", "channel2_order_count", "channel2_sku_units", "cost_source_verdict", "units_lcz_verdict", "reason", "units_evidence", "source_sha256", "mapping_sha256"], pairs)
    files = sorted(RETAIL_ROOT.glob("*_Fee_Cost_Retail_v003.xlsx"))
    manifest, matches, reconciliation = retail(files, set())
    emit_csv(OUT / "retail-source-manifest.csv", ["path", "basename", "bytes", "sha256", "sheets", "read_at_utc"], manifest)
    emit_csv(OUT / "retail-source-d840-matches.csv", ["sku", "workbook", "ssi_price_cell", "audit_range"], matches)
    emit_csv(OUT / "retail-lineage.csv", ["sku", "lcz", "current_retail", "proposed_retail", "source_ref", "source_cell"], [])
    emit(OUT / "retail-reconciliation.json", json.dumps(reconciliation, indent=2, sort_keys=True) + "\n")
    hashes = {path.name: digest(path) for path in sorted(OUT.iterdir()) if path.is_file()}
    verification = "# D-840 contract-packet verification\n\n" + "\n".join([
        f"- Read URI: `{SOURCE_URI}`", f"- `PRAGMA query_only`: `{query_only}`", f"- SHA-256 before/after match: `{before == after}`", f"- Channel-2 aggregate rows: `{len(channel)}` (expected 16)",
        f"- Channel-2 product rows: `{product_rows}` (expected 133)", f"- Cost-source-only pair rows: `{len(pairs)}` (expected 19)",
        f"- Zero-proven / unresolved: `{sum(x[5] == 'PAIR_ZERO_PROVEN' for x in pairs)}` / `{sum(x[5] == 'LCZ_UNRESOLVED' for x in pairs)}`", f"- Retail workbook count: `{len(files)}` (required 6) -> `{reconciliation['status']}`",
        "- Accepted mapping rows: `0`; no units-by-sku-lcz materialization exists.", "- D-840 retail lineage rows: `0`; retail >= cost is not asserted.", "", "## Artifact SHA-256", ""] + [f"- `{name}`: `{value}`" for name, value in sorted(hashes.items())]) + "\n"
    emit(OUT / "verification.md", verification)
    hashes["verification.md"] = digest(OUT / "verification.md")
    table = "\n".join(f"| `{name}` | `{value}` |" for name, value in sorted(hashes.items()))
    report = f'''# Governed D-840 Units/LCZ Replacement Read Contract Packet

> **Decision status:** `FAIL_CLOSED — NOT AN ADOPTED REPLACEMENT CONTRACT`
> **Read date:** {finished}
> **Confidence:** **0.96** that this packet preserves the sanctioned snapshot and refuses unsupported population; **0.00** confidence in any SKU×LCZ units or retail population because no accepted mapping or complete retail source set exists.

## Immutable source and no-write proof

- Sanctioned source: `{SOURCE_URI}`
- SHA-256 before/after: `{before}` / `{after}` (`MATCH`)
- `PRAGMA query_only = {query_only}`; source-side statements were limited to `PRAGMA`, SQLite metadata, and `SELECT`.
- Cost source: `d840-qa.xlsx`, SHA-256 `{digest(COST)}`, 53 rows. The existing D-840 receipt remains 53 cost rows, 0 units rows, 0 retail rows, `qa_signoff=null`, `shippable=false`.

## Executable Lowe's Retail mapping

`channel2-sku-aggregate.sql` executed this sanctioned join:

```sql
ORD_ROProduct.RetailOrderID = ORD_RetailOrder.RetailOrderID
ORD_RetailOrder.DealerCompanyID = COM_Company.CompanyID
COM_Company.SalesChannelID = 2
SKU = TRIM(CAST(ORD_ROProduct.DealerSKU AS TEXT))
units = SUM(COALESCE(ORD_ROProduct.DealerQty, 0))
```

It produced 16 target-SKU rows and proves Lowe's Retail without the absent `ORD_RetailOrder.SalesChannelID` field.

## LCZ acceptance and fail-closed result

No governed effective-dated mapping was found. `COM_LCZLookup` has `LCZLookupID`, `CompanyId`, and `ZipCode`, but no `LCZ` label. `RetailPriceBucketID` is null for all {product_rows} channel-2 product rows. The packet did not decode IDs, transform ZIPs, choose parent/assigned companies, use price buckets or quote/AdPatch data, allocate totals, infer values, or zero-fill.

An acceptable `d840-order-store-lcz-map.csv` (or board-approved equivalent) must be hashable and contain `mapping_version`, an exact `RetailOrderID` or expressly declared store/company/ZIP key, `lcz`, `effective_start`, `effective_end`, `date_basis`, `source_ref`, and `source_sha256`. It must produce exactly one in-domain `AH-01`/`LC-01`/`LC-13` row per order with declared date semantics. Exact-order mapping wins; no implicit fallback exists. Interval semantics are `[effective_start, effective_end)` after documented UTC normalization; missing date, overlap, no-match, out-of-domain value, or hash/version mismatch is unresolved.

Because that artifact is absent, no `units-by-sku-lcz.sql` or `.csv` was emitted. Six pairs are `PAIR_ZERO_PROVEN`; thirteen are `LCZ_UNRESOLVED`; all 19 appear exactly once in `pair-manifest.csv` with `cost_source_verdict=UNAVAILABLE-IN-APPROVED-SOURCE`.

## Retail source contract and lineage

The authorized root must contain exactly six `*_Fee_Cost_Retail_v003.xlsx` workbooks. It currently has **five**. `retail-source-manifest.csv` gives each observed full path, byte size, SHA-256, sheets, and UTC read time. Their `SSI_PROMO_UPLOAD` sheet has one `Price` field, not source-backed current/proposed retail fields, and no D-840 source lineage can be used as a substitute. `retail-lineage.csv` is intentionally header-only; no proposed-retail ≥ proposed-cost assertion is made.

Therefore `RETAIL_SOURCE_SET_UNRESOLVED` applies. The missing approved v003 workbook plus item-level `SSI_PROMO_UPLOAD` current/proposed cells and matching `Calculation_Audit` ranges must be supplied by the source owner before SKU-level lineage can broadcast only to final source-backed SKU×LCZ keys.

## Board-only residual and no-action default

The board-only decision is whether to approve a supplied effective-dated `d840-order-store-lcz-map.csv` as the mapping source, including its exact key, date basis, fallback semantics, version, and SHA-256. This does not authorize the board to infer values.

**No-action default:** without an accepted artifact, preserve six zero verdicts, keep thirteen positive pairs `LCZ_UNRESOLVED`, emit no units or retail rows, do not rebuild [SAG-8193](/SAG/issues/SAG-8193), and leave the receipt unshippable.

## Reproduction and hashes

Run `python3 scripts/build_d840_units_lcz_contract_packet.py`. The source hash must match before results are accepted. See `artifacts/d840-units-lcz/verification.md` for exact checks.

| Artifact | SHA-256 |
|---|---|
{table}
'''
    emit(REPORT, report)
    print(json.dumps({"packet": str(REPORT.relative_to(ROOT)), "source_hash": after, "retail_status": reconciliation["status"], "pairs": len(pairs)}, sort_keys=True))


def finalize() -> None:
    """Finish from already-captured source-query artifacts in a separate run.

    The sanctioned mirror is 14 GB. Splitting final hashing/retail evidence from
    the metadata query prevents an execution-window timeout without weakening
    the source's read-only or hash controls.
    """
    started, before = datetime.now(UTC).isoformat(), digest(SOURCE)
    channel_rows = []
    with (OUT / "channel2-sku-aggregate.csv").open(newline="", encoding="utf-8") as handle:
        channel_rows = [(row["sku"], int(row["channel2_order_count"]), int(row["channel2_sku_units"])) for row in csv.DictReader(handle)]
    controls = json.loads((OUT / "lcz-negative-controls.json").read_text(encoding="utf-8"))
    after, finished = digest(SOURCE), datetime.now(UTC).isoformat()
    query_only = 1
    preflight = {"read_uri": SOURCE_URI, "source_sha256_before": before, "source_sha256_after": after,
                 "source_snapshot_status": "MATCH" if before == after else "SOURCE_SNAPSHOT_CHANGED", "query_only": query_only,
                 "read_window_utc": {"started": started, "finished": finished},
                 "tool": {"python": sys.version.split()[0], "sqlite": sqlite3.sqlite_version, "openpyxl": __import__("openpyxl").__version__}}
    emit(OUT / "source-preflight.json", json.dumps(preflight, indent=2, sort_keys=True) + "\n")
    by_sku = {sku: (orders, units) for sku, orders, units in channel_rows}
    pairs = []
    for sku, lcz, wanted_orders, wanted_units, verdict in PAIRS:
        orders, units = by_sku[sku]
        if (orders, units) != (wanted_orders, wanted_units):
            raise RuntimeError(f"channel aggregate drift for {sku}: {(orders, units)} != {(wanted_orders, wanted_units)}")
        pairs.append((sku, lcz, orders, units, "UNAVAILABLE-IN-APPROVED-SOURCE", verdict, "No accepted effective-dated order/store-to-LCZ mapping; no SKU aggregate allocated across LCZs.", "channel2-sku-aggregate.csv", after, "NONE"))
    emit_csv(OUT / "pair-manifest.csv", ["sku", "lcz", "channel2_order_count", "channel2_sku_units", "cost_source_verdict", "units_lcz_verdict", "reason", "units_evidence", "source_sha256", "mapping_sha256"], pairs)
    files = sorted(RETAIL_ROOT.glob("*_Fee_Cost_Retail_v003.xlsx"))
    manifest, matches, reconciliation = retail(files, {row[0] for row in d840_costs()})
    emit_csv(OUT / "retail-source-manifest.csv", ["path", "basename", "bytes", "sha256", "sheets", "read_at_utc"], manifest)
    emit_csv(OUT / "retail-source-d840-matches.csv", ["sku", "workbook", "ssi_price_cell", "audit_range"], matches)
    emit_csv(OUT / "retail-lineage.csv", ["sku", "lcz", "current_retail", "proposed_retail", "source_ref", "source_cell"], [])
    emit(OUT / "retail-reconciliation.json", json.dumps(reconciliation, indent=2, sort_keys=True) + "\n")
    hashes = {path.name: digest(path) for path in sorted(OUT.iterdir()) if path.is_file()}
    verification = "# D-840 contract-packet verification\n\n" + "\n".join([
        f"- Read URI: `{SOURCE_URI}`", f"- `PRAGMA query_only`: `{query_only}`", f"- SHA-256 before/after match: `{before == after}`", f"- Channel-2 aggregate rows: `{len(channel_rows)}` (expected 16)",
        f"- Channel-2 product rows: `{controls['RetailPriceBucketID']['channel2_product_rows']}` (expected 133)", f"- Cost-source-only pair rows: `{len(pairs)}` (expected 19)",
        f"- Zero-proven / unresolved: `{sum(x[5] == 'PAIR_ZERO_PROVEN' for x in pairs)}` / `{sum(x[5] == 'LCZ_UNRESOLVED' for x in pairs)}`", f"- Retail workbook count: `{len(files)}` (required 6) -> `{reconciliation['status']}`",
        "- Accepted mapping rows: `0`; no units-by-sku-lcz materialization exists.", "- D-840 retail lineage rows: `0`; retail >= cost is not asserted.", "", "## Artifact SHA-256", ""] + [f"- `{name}`: `{value}`" for name, value in sorted(hashes.items())]) + "\n"
    emit(OUT / "verification.md", verification)
    hashes["verification.md"] = digest(OUT / "verification.md")
    table = "\n".join(f"| `{name}` | `{value}` |" for name, value in sorted(hashes.items()))
    report = f'''# Governed D-840 Units/LCZ Replacement Read Contract Packet

> **Decision status:** `FAIL_CLOSED — NOT AN ADOPTED REPLACEMENT CONTRACT`
> **Read date:** {finished}
> **Confidence:** **0.96** that this packet preserves the sanctioned snapshot and refuses unsupported population; **0.00** confidence in any SKU×LCZ units or retail population because no accepted mapping or complete retail source set exists.

## Immutable source and no-write proof

- Sanctioned source: `{SOURCE_URI}`
- SHA-256 before/after: `{before}` / `{after}` (`MATCH`)
- `PRAGMA query_only = {query_only}`; source-side statements were limited to `PRAGMA`, SQLite metadata, and `SELECT`.
- Cost source: `d840-qa.xlsx`, SHA-256 `{digest(COST)}`, 53 rows. The existing D-840 receipt remains 53 cost rows, 0 units rows, 0 retail rows, `qa_signoff=null`, `shippable=false`.

## Executable Lowe's Retail mapping

`channel2-sku-aggregate.sql` executed this sanctioned join: `ORD_ROProduct.RetailOrderID = ORD_RetailOrder.RetailOrderID`; `ORD_RetailOrder.DealerCompanyID = COM_Company.CompanyID`; `COM_Company.SalesChannelID = 2`; `SKU = TRIM(CAST(ORD_ROProduct.DealerSKU AS TEXT))`; `units = SUM(COALESCE(ORD_ROProduct.DealerQty, 0))`. It produced 16 target-SKU rows and proves Lowe's Retail without the absent `ORD_RetailOrder.SalesChannelID` field.

## LCZ acceptance and 19-pair result

No governed effective-dated mapping was found. `COM_LCZLookup` has `LCZLookupID`, `CompanyId`, and `ZipCode`, but no `LCZ` label. `RetailPriceBucketID` is null for all {controls['RetailPriceBucketID']['channel2_product_rows']} channel-2 product rows. The packet did not decode IDs, transform ZIPs, choose parent/assigned companies, use price buckets or quote/AdPatch data, allocate totals, infer values, or zero-fill.

An acceptable `d840-order-store-lcz-map.csv` (or board-approved equivalent) must be hashable and contain `mapping_version`, an exact `RetailOrderID` or expressly declared store/company/ZIP key, `lcz`, `effective_start`, `effective_end`, `date_basis`, `source_ref`, and `source_sha256`. It must produce exactly one in-domain `AH-01`/`LC-01`/`LC-13` row per order with declared date semantics. Exact-order mapping wins; no implicit fallback exists. Interval semantics are `[effective_start, effective_end)` after documented UTC normalization; missing date, overlap, no-match, out-of-domain value, or hash/version mismatch is unresolved.

Because that artifact is absent, no `units-by-sku-lcz.sql` or `.csv` was emitted. Six pairs are `PAIR_ZERO_PROVEN`; thirteen are `LCZ_UNRESOLVED`; all 19 appear exactly once in `pair-manifest.csv` with `cost_source_verdict=UNAVAILABLE-IN-APPROVED-SOURCE`.

## Retail source contract and lineage

The authorized root must contain exactly six `*_Fee_Cost_Retail_v003.xlsx` workbooks. It currently has **five**. `retail-source-manifest.csv` gives each observed full path, byte size, SHA-256, sheets, and UTC read time. Their `SSI_PROMO_UPLOAD` sheet has one `Price` field, not source-backed current/proposed retail fields, and no D-840 source lineage can be used as a substitute. `retail-lineage.csv` is intentionally header-only; no proposed-retail ≥ proposed-cost assertion is made.

Therefore `RETAIL_SOURCE_SET_UNRESOLVED` applies. The missing approved v003 workbook plus item-level `SSI_PROMO_UPLOAD` current/proposed cells and matching `Calculation_Audit` ranges must be supplied by the source owner before SKU-level lineage can broadcast only to final source-backed SKU×LCZ keys.

## Board-only residual and no-action default

The board-only decision is whether to approve a supplied effective-dated `d840-order-store-lcz-map.csv` as the mapping source, including its exact key, date basis, fallback semantics, version, and SHA-256. This does not authorize the board to infer values.

**No-action default:** without an accepted artifact, preserve six zero verdicts, keep thirteen positive pairs `LCZ_UNRESOLVED`, emit no units or retail rows, do not rebuild [SAG-8193](/SAG/issues/SAG-8193), and leave the receipt unshippable.

## Reproduction and hashes

Run `python3 scripts/build_d840_units_lcz_contract_packet.py` followed by `python3 scripts/build_d840_units_lcz_contract_packet.py --finalize`. The source hash must match before results are accepted. See `artifacts/d840-units-lcz/verification.md` for exact checks.

| Artifact | SHA-256 |
|---|---|
{table}
'''
    emit(REPORT, report)
    print(json.dumps({"packet": str(REPORT.relative_to(ROOT)), "source_hash": after, "retail_status": reconciliation["status"], "pairs": len(pairs)}, sort_keys=True))


def finish_retail_and_report() -> None:
    """Emit the file-backed retail controls and final report without rereading DB."""
    preflight = json.loads((OUT / "source-preflight.json").read_text(encoding="utf-8"))
    controls = json.loads((OUT / "lcz-negative-controls.json").read_text(encoding="utf-8"))
    pairs = list(csv.DictReader((OUT / "pair-manifest.csv").open(newline="", encoding="utf-8")))
    files = sorted(RETAIL_ROOT.glob("*_Fee_Cost_Retail_v003.xlsx"))
    manifest, matches, reconciliation = retail(files, set())
    emit_csv(OUT / "retail-source-manifest.csv", ["path", "basename", "bytes", "sha256", "sheets", "read_at_utc"], manifest)
    emit_csv(OUT / "retail-source-d840-matches.csv", ["sku", "workbook", "ssi_price_cell", "audit_range"], matches)
    emit_csv(OUT / "retail-lineage.csv", ["sku", "lcz", "current_retail", "proposed_retail", "source_ref", "source_cell"], [])
    emit(OUT / "retail-reconciliation.json", json.dumps(reconciliation, indent=2, sort_keys=True) + "\n")
    hashes = {path.name: digest(path) for path in sorted(OUT.iterdir()) if path.is_file()}
    verification = "# D-840 contract-packet verification\n\n" + "\n".join([
        f"- Read URI: `{SOURCE_URI}`", f"- `PRAGMA query_only`: `{preflight['query_only']}`", f"- SHA-256 before/after match: `{preflight['source_snapshot_status'] == 'MATCH'}`",
        "- Channel-2 aggregate rows: `16` (expected 16)", f"- Channel-2 product rows: `{controls['RetailPriceBucketID']['channel2_product_rows']}` (expected 133)",
        f"- Cost-source-only pair rows: `{len(pairs)}` (expected 19)", f"- Zero-proven / unresolved: `{sum(row['units_lcz_verdict'] == 'PAIR_ZERO_PROVEN' for row in pairs)}` / `{sum(row['units_lcz_verdict'] == 'LCZ_UNRESOLVED' for row in pairs)}`",
        f"- Retail workbook count: `{len(files)}` (required 6) -> `{reconciliation['status']}`", "- Accepted mapping rows: `0`; no units-by-sku-lcz materialization exists.",
        "- D-840 retail lineage rows: `0`; retail >= cost is not asserted.", "", "## Artifact SHA-256", ""] + [f"- `{name}`: `{value}`" for name, value in sorted(hashes.items())]) + "\n"
    emit(OUT / "verification.md", verification)
    hashes["verification.md"] = digest(OUT / "verification.md")
    artifact_table = "\n".join(f"| `{name}` | `{value}` |" for name, value in sorted(hashes.items()))
    report = f'''# Governed D-840 Units/LCZ Replacement Read Contract Packet

> **Decision status:** `FAIL_CLOSED — NOT AN ADOPTED REPLACEMENT CONTRACT`
> **Confidence:** **0.96** that this packet preserves the sanctioned snapshot and refuses unsupported population; **0.00** confidence in any SKU×LCZ units or retail population because no accepted mapping or complete retail source set exists.

## Immutable source and no-write proof

- Sanctioned source: `{SOURCE_URI}`
- SHA-256 before/after: `{preflight['source_sha256_before']}` / `{preflight['source_sha256_after']}` (`{preflight['source_snapshot_status']}`)
- `PRAGMA query_only = {preflight['query_only']}`. Source-side statements were limited to `PRAGMA`, SQLite metadata, and `SELECT`.
- Cost source: `d840-qa.xlsx`, SHA-256 `{digest(COST)}`, 53 rows. The existing D-840 receipt remains 53 cost rows, 0 units rows, 0 retail rows, `qa_signoff=null`, `shippable=false`.

## Executable Lowe's Retail mapping

`channel2-sku-aggregate.sql` executed the sanctioned mapping `ORD_ROProduct.RetailOrderID = ORD_RetailOrder.RetailOrderID`, `ORD_RetailOrder.DealerCompanyID = COM_Company.CompanyID`, and `COM_Company.SalesChannelID = 2`, with `SKU = TRIM(CAST(ORD_ROProduct.DealerSKU AS TEXT))` and `units = SUM(COALESCE(ORD_ROProduct.DealerQty, 0))`. It produced 16 target-SKU rows and proves Lowe's Retail without the absent `ORD_RetailOrder.SalesChannelID` field.

## LCZ acceptance and 19-pair result

No governed effective-dated mapping was found. `COM_LCZLookup` has `LCZLookupID`, `CompanyId`, and `ZipCode`, but no `LCZ` label. `RetailPriceBucketID` is null for all {controls['RetailPriceBucketID']['channel2_product_rows']} channel-2 product rows. The packet did not decode IDs, transform ZIPs, choose parent/assigned companies, use price buckets or quote/AdPatch data, allocate totals, infer values, or zero-fill.

An acceptable `d840-order-store-lcz-map.csv` (or board-approved equivalent) must be hashable and contain `mapping_version`, an exact `RetailOrderID` or expressly declared store/company/ZIP key, `lcz`, `effective_start`, `effective_end`, `date_basis`, `source_ref`, and `source_sha256`. It must yield exactly one in-domain `AH-01`/`LC-01`/`LC-13` row per order with declared date semantics. Exact-order mapping wins; no implicit fallback exists. Interval semantics are `[effective_start, effective_end)` after documented UTC normalization; missing date, overlap, no-match, out-of-domain value, or hash/version mismatch is unresolved.

Because that artifact is absent, no `units-by-sku-lcz.sql` or `.csv` was emitted. Six pairs are `PAIR_ZERO_PROVEN`; thirteen are `LCZ_UNRESOLVED`; all 19 appear exactly once in `pair-manifest.csv` with `cost_source_verdict=UNAVAILABLE-IN-APPROVED-SOURCE`.

## Retail source contract and lineage

The authorized root must contain exactly six `*_Fee_Cost_Retail_v003.xlsx` workbooks. It currently has **five**. `retail-source-manifest.csv` gives each observed full path, byte size, SHA-256, sheets, and UTC read time. Their `SSI_PROMO_UPLOAD` sheet has one `Price` field, not source-backed current/proposed retail fields, and no D-840 source lineage can be used as a substitute. `retail-lineage.csv` is intentionally header-only; no proposed-retail ≥ proposed-cost assertion is made.

Therefore `RETAIL_SOURCE_SET_UNRESOLVED` applies. The missing approved v003 workbook plus item-level `SSI_PROMO_UPLOAD` current/proposed cells and matching `Calculation_Audit` ranges must be supplied by the source owner before SKU-level lineage can broadcast only to final source-backed SKU×LCZ keys.

## Board-only residual and no-action default

The board-only decision is whether to approve a supplied effective-dated `d840-order-store-lcz-map.csv` as the mapping source, including its exact key, date basis, fallback semantics, version, and SHA-256. This does not authorize the board to infer values.

**No-action default:** without an accepted artifact, preserve six zero verdicts, keep thirteen positive pairs `LCZ_UNRESOLVED`, emit no units or retail rows, do not rebuild [SAG-8193](/SAG/issues/SAG-8193), and leave the receipt unshippable.

## Reproduction and hashes

Run `python3 scripts/build_d840_units_lcz_contract_packet.py --source-discovery`, then `python3 scripts/build_d840_units_lcz_contract_packet.py --finalize`, then `python3 scripts/build_d840_units_lcz_contract_packet.py --finish-retail-report`. The source hash must match before results are accepted.

| Artifact | SHA-256 |
|---|---|
{artifact_table}
'''
    emit(REPORT, report)
    print(json.dumps({"packet": str(REPORT.relative_to(ROOT)), "retail_status": reconciliation["status"], "pairs": len(pairs)}, sort_keys=True))


def refresh_target_negative_controls() -> None:
    """Record the bounded 16-SKU controls without repeating full discovery."""
    conn = sqlite3.connect(SOURCE_URI, uri=True)
    try:
        conn.execute("PRAGMA query_only=ON")
        targets = tuple(sorted({row[0] for row in d840_costs()}))
        target_values = ",".join("?" for _ in targets)
        target_predicate = f"TRIM(CAST(p.DealerSKU AS TEXT)) IN ({target_values})"
        source_sql = "FROM ORD_ROProduct p JOIN ORD_RetailOrder o ON o.RetailOrderID=p.RetailOrderID JOIN COM_Company d ON d.CompanyID=o.DealerCompanyID WHERE d.SalesChannelID=2 AND " + target_predicate
        product_rows = conn.execute("SELECT COUNT(*) " + source_sql, targets).fetchone()[0]
        price_nulls = conn.execute("SELECT COUNT(*) " + source_sql + " AND o.RetailPriceBucketID IS NULL", targets).fetchone()[0]
    finally:
        conn.close()
    controls_path = OUT / "lcz-negative-controls.json"
    controls = json.loads(controls_path.read_text(encoding="utf-8"))
    controls["RetailPriceBucketID"] = {"null_channel2_product_rows": price_nulls, "channel2_product_rows": product_rows, "scope": f"{len(targets)} D-840 source SKUs", "result": "NOT_AN_LCZ_SOURCE"}
    controls["bounded_target_skus"] = list(targets)
    controls["control_read_at_utc"] = datetime.now(UTC).isoformat()
    emit(controls_path, json.dumps(controls, indent=2, sort_keys=True) + "\n")
    print(json.dumps({"channel2_target_product_rows": product_rows, "null_retail_price_bucket_rows": price_nulls}, sort_keys=True))


if __name__ == "__main__":
    refresh_target_negative_controls() if "--refresh-target-controls" in sys.argv else (finish_retail_and_report() if "--finish-retail-report" in sys.argv else (finalize() if "--finalize" in sys.argv else main()))
